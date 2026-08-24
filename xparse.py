import openpyxl, sys

def dump(fn, tabs):
    wb = openpyxl.load_workbook(fn, read_only=True, data_only=True)
    print(f"\n########## {fn} ##########")
    for tab in tabs:
        if tab not in wb.sheetnames:
            print(f"== {tab}: MISSING")
            continue
        ws = wb[tab]
        rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None and str(c).strip()!='' for c in r)]
        print(f"\n=== {tab}: {len(rows)} rows ===")
        for i, r in enumerate(rows):
            vals = [str(c)[:30] if c is not None else '·' for c in r[:14]]
            print(f"R{i}: " + " | ".join(vals))

dump(sys.argv[1], ["Inventory", "Sales_Aug_2026", "Wastage_Ledger", "Returns", "Stock_Adjustments", "Batches"])
