#!/usr/bin/env python3
"""Tillzo sheet verify: latest xlsx export -> key tabs summary. Usage: python verify_sheet.py [tab1 tab2 ...]"""
import sys, glob, os
import openpyxl

dl = r"C:\Users\Faisal Khan\Downloads"
# pick newest matching export
cands = glob.glob(os.path.join(dl, "Faisal Mart — TillzoPOS*.xlsx"))
if not cands:
    print("NO XLSX FOUND"); sys.exit(1)
fn = max(cands, key=os.path.getmtime)
print("FILE:", os.path.basename(fn), os.path.getmtime(fn))
wb = openpyxl.load_workbook(fn, read_only=True, data_only=True)
print("TABS:", wb.sheetnames)
tabs = sys.argv[1:] or ["Customers","Vendors","Expenses","Inventory","Sales_Aug_2026","Wastage_Ledger","Stock_Adjustments","Till_Sessions","Purchase_Orders","GRN_Headers","Returns","Khata_Events","Categories","Product_Units","Time_Clock"]
for tab in tabs:
    if tab not in wb.sheetnames:
        print(f"\n=== {tab}: MISSING TAB ==="); continue
    ws = wb[tab]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None and str(c).strip()!='' for c in r)]
    print(f"\n=== {tab}: {len(rows)} non-empty rows ===")
    for r in rows[:8]:
        print(" | ".join((str(c)[:24] if c is not None else '·') for c in r[:12]))
