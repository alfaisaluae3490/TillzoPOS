import openpyxl, sys
fn = r"C:\Users\Faisal Khan\Downloads\Faisal Mart — TillzoPOS (1).xlsx"
wb = openpyxl.load_workbook(fn, read_only=True, data_only=True)
print("TABS:", wb.sheetnames)
for ws in wb.worksheets:
    rows = list(ws.iter_rows(values_only=True))
    nonempty = [r for r in rows if any(c is not None and str(c).strip() != '' for c in r)]
    print(f"\n=== {ws.title}: {len(nonempty)} non-empty rows ===")
    if nonempty:
        hdr = nonempty[0]
        print("HDR:", [str(c)[:22] if c is not None else '' for c in hdr][:18])
